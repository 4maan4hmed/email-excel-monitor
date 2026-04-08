#!/usr/bin/env python3
import imaplib
import email
from email.header import decode_header
import os
import sys
from datetime import datetime, timedelta
import requests
import openpyxl
import io
import html
import json

PROCESSED_FILE = 'processed_emails.json'

def load_processed_emails():
    if os.path.exists(PROCESSED_FILE):
        try:
            with open(PROCESSED_FILE, 'r') as f:
                return set(json.load(f))
        except Exception as e:
            print(f"Error loading processed emails: {e}")
            return set()
    return set()

def save_processed_emails(processed_set):
    with open(PROCESSED_FILE, 'w') as f:
        json.dump(list(processed_set), f)

def escape_html(text):
    if text:
        return html.escape(str(text))
    return ""

def send_telegram_message(bot_token, chat_id, message):
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    if len(message) > 4090:
        message = message[:4000] + "\n\n[Message truncated]"
    
    payload = {'chat_id': chat_id, 'text': message, 'parse_mode': 'HTML'}
    try:
        response = requests.post(url, json=payload, timeout=10)
        response.raise_for_status()
        print(f"✓ Telegram sent")
        return True
    except Exception as e:
        print(f"Telegram send error: {e}")
        try:
            payload['text'] = message.replace('<b>', '').replace('</b>', '').replace('<pre>', '').replace('</pre>', '')
            payload.pop('parse_mode')
            requests.post(url, json=payload, timeout=10)
            print(f"✓ Telegram sent (without HTML)")
        except Exception as e2:
            print(f"✗ Failed to send Telegram: {e2}")
        return False

def scan_excel_for_name(file_data, search_name, filename):
    matches = []
    try:
        print(f"    Scanning Excel file: {filename}")
        workbook = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
        print(f"    Sheets found: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"    Checking sheet: {sheet_name} (rows: {sheet.max_row}, cols: {sheet.max_column})")
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value and search_name.lower() in str(cell_value).lower():
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        matches.append({
                            'sheet': sheet_name,
                            'cell': f"{col_letter}{row_idx}",
                            'value': str(cell_value)[:100]  # Limit value length
                        })
                        print(f"      Match found at {sheet_name}!{col_letter}{row_idx}: {str(cell_value)[:50]}")
        
        workbook.close()
    except Exception as e:
        print(f"    ✗ Error scanning {filename}: {e}")
    return matches

def decode_email_subject(subject):
    if subject is None:
        return "No Subject"
    decoded_parts = []
    for part, encoding in decode_header(subject):
        if isinstance(part, bytes):
            try:
                decoded_parts.append(part.decode(encoding or 'utf-8', errors='ignore'))
            except:
                decoded_parts.append(part.decode('utf-8', errors='ignore'))
        else:
            decoded_parts.append(str(part))
    return ''.join(decoded_parts)

def get_email_body(msg):
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain" and "attachment" not in str(part.get_content_disposition()):
                try:
                    body = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                    break
                except:
                    try:
                        body = part.get_payload(decode=True).decode('latin-1', errors='ignore')
                        break
                    except:
                        pass
    else:
        try:
            body = msg.get_payload(decode=True).decode('utf-8', errors='ignore')
        except:
            try:
                body = msg.get_payload(decode=True).decode('latin-1', errors='ignore')
            except:
                pass
    return body if body else ""

def connect_to_email(email_address, app_password, imap_server='imap.gmail.com'):
    try:
        mail = imaplib.IMAP4_SSL(imap_server)
        mail.login(email_address, app_password)
        print(f"✓ Logged in as {email_address}")
        return mail
    except Exception as e:
        print(f"✗ Connection error: {e}")
        sys.exit(1)

def process_emails(mail, search_name, bot_token, chat_id, check_last_minutes):
    processed_emails = load_processed_emails()
    print(f"Already processed: {len(processed_emails)} emails")
    
    mail.select('INBOX')
    
    since_time = datetime.now() - timedelta(minutes=check_last_minutes)
    since_date = since_time.strftime("%d-%b-%Y")
    
    print(f"Checking emails since {since_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Searching for term: '{search_name}' (case-insensitive)")
    
    status, messages = mail.search(None, f'SINCE {since_date}')
    if status != 'OK':
        print("Failed to search emails")
        return
    
    email_ids = messages[0].split()
    if not email_ids:
        print(f"No emails in last {check_last_minutes} minutes")
        return
    
    print(f"Found {len(email_ids)} email(s) in time window")
    
    matches_found = 0
    for email_id in email_ids:
        try:
            email_id_str = email_id.decode()
            
            if email_id_str in processed_emails:
                print(f"\n⏭ Skipping already processed email ID: {email_id_str}")
                continue
            
            status, msg_data = mail.fetch(email_id, '(RFC822)')
            if status != 'OK':
                print(f"Failed to fetch email {email_id_str}")
                continue
            
            raw_email = msg_data[0][1]
            email_message = email.message_from_bytes(raw_email)
            
            subject = decode_email_subject(email_message.get('Subject'))
            from_email = email_message.get('From')
            date = email_message.get('Date')
            
            # Check date
            try:
                from email.utils import parsedate_to_datetime
                email_datetime = parsedate_to_datetime(date)
                if email_datetime < since_time:
                    print(f"\n⏭ Skipping old email: {subject[:50]}")
                    processed_emails.add(email_id_str)
                    continue
            except Exception as e:
                print(f"Date parse error: {e}")
            
            print(f"\n📧 Processing: {subject[:80]}")
            print(f"   From: {from_email}")
            print(f"   Date: {date}")
            
            # Get body
            body = get_email_body(email_message)
            body_has_match = search_name.lower() in body.lower() if body else False
            
            if body:
                print(f"   Body length: {len(body)} chars")
                if body_has_match:
                    print(f"   ✓ Body contains '{search_name}'")
            else:
                print(f"   No text body found")
            
            # Check attachments
            excel_matches = []
            attachment_count = 0
            
            for part in email_message.walk():
                if part.get_content_disposition() == 'attachment':
                    filename = part.get_filename()
                    if filename:
                        filename = decode_email_subject(filename)
                        attachment_count += 1
                        print(f"   Attachment {attachment_count}: {filename}")
                        
                        if filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                            print(f"   📊 Scanning Excel: {filename}")
                            file_data = part.get_payload(decode=True)
                            print(f"   File size: {len(file_data)} bytes")
                            matches = scan_excel_for_name(file_data, search_name, filename)
                            if matches:
                                excel_matches.extend(matches)
                                print(f"   ✓ Found {len(matches)} match(es) in Excel")
                            else:
                                print(f"   No matches in Excel")
            
            if attachment_count == 0:
                print(f"   No attachments found")
            
            # Send notification if match found
            if body_has_match or excel_matches:
                matches_found += 1
                print(f"  🎯 MATCH FOUND! (#{matches_found})")
                
                safe_search = escape_html(search_name)
                safe_subject = escape_html(subject)
                safe_from = escape_html(from_email)
                safe_date = escape_html(date)
                safe_body = escape_html(" ".join(body.split())[:500])
                
                message = f"<b>🔔 MATCH: {safe_search}</b>\n────────────────────\n"
                
                if excel_matches:
                    message += f"📊 Excel: {len(excel_matches)} match(es)\n"
                if body_has_match:
                    message += f"📧 Email body: Match found\n"
                
                message += (
                    f"\n<b>EMAIL</b>\n"
                    f"Subject: {safe_subject}\n"
                    f"From: {safe_from}\n"
                    f"Date: {safe_date}\n"
                )
                
                if excel_matches:
                    message += f"\n<b>EXCEL MATCHES</b>\n"
                    for i, match in enumerate(excel_matches[:5], 1):
                        safe_val = escape_html(match['value'][:50])
                        message += f"{i}. {match['sheet']} ({match['cell']}): {safe_val}\n"
                    if len(excel_matches) > 5:
                        message += f"... +{len(excel_matches) - 5} more\n"
                
                if body_has_match:
                    message += f"\n<b>BODY PREVIEW</b>\n<pre>{safe_body}</pre>"
                
                send_telegram_message(bot_token, chat_id, message)
                processed_emails.add(email_id_str)
            else:
                print(f"  ❌ No match found for '{search_name}'")
                processed_emails.add(email_id_str)
                
        except Exception as e:
            print(f"✗ Error processing email {email_id}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    save_processed_emails(processed_emails)
    print(f"\n📊 Summary: Found {matches_found} matching email(s) out of {len(email_ids)} checked")

def main():
    # Load environment variables
    email_address = os.getenv('EMAIL_ADDRESS')
    app_password = os.getenv('EMAIL_APP_PASSWORD')
    search_name = os.getenv('SEARCH_NAME')
    bot_token = os.getenv('TELEGRAM_BOT_TOKEN')
    chat_id = os.getenv('TELEGRAM_CHAT_ID')
    imap_server = os.getenv('IMAP_SERVER', 'imap.gmail.com')
    
    # Debug: Print environment (mask sensitive data)
    print("--- Environment Check ---")
    print(f"EMAIL_ADDRESS: {email_address if email_address else '❌ MISSING'}")
    print(f"EMAIL_APP_PASSWORD: {'✓ SET' if app_password else '❌ MISSING'}")
    print(f"SEARCH_NAME: '{search_name}'" if search_name else "SEARCH_NAME: ❌ MISSING")
    print(f"TELEGRAM_BOT_TOKEN: {'✓ SET' if bot_token else '❌ MISSING'}")
    print(f"TELEGRAM_CHAT_ID: {'✓ SET' if chat_id else '❌ MISSING'}")
    print(f"IMAP_SERVER: {imap_server}")
    print("------------------------")
    
    if not all([email_address, app_password, search_name, bot_token, chat_id]):
        print("✗ Missing required environment variables")
        print("Required: EMAIL_ADDRESS, EMAIL_APP_PASSWORD, SEARCH_NAME, TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID")
        sys.exit(1)
    
    try:
        check_last_minutes = int(os.getenv('CHECK_LAST_MINUTES', '90'))
    except:
        check_last_minutes = 90
    
    print(f"\n--- Started {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---")
    print(f"Check window: {check_last_minutes} minutes")
    
    mail = connect_to_email(email_address, app_password, imap_server)
    
    try:
        process_emails(mail, search_name, bot_token, chat_id, check_last_minutes)
    except Exception as e:
        print(f"Fatal error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            mail.logout()
            print("\n✓ Disconnected from email server")
        except:
            pass
    
    print(f"--- Finished {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---")

if __name__ == "__main__":
    main()